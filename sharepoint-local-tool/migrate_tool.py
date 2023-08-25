# Import necessary libraries
import os
import distutils.dir_util
from openpyxl import Workbook, load_workbook

# Define global variables
content = []
options = []
new_path = ""


# Define function for content reading
def read_content(path: str) -> list:
    """
    Return the list of subdirectiories name from a given path of the directory
    :param path: path of the desired directory in string format
    :return: content: list of subdirectories name in string format
    """
    # Modify global variable
    global content

    # Retrieve subdirectories name
    content = os.listdir(path)
    
    # Return modified content variable
    return content


# Define function for path generation
def generate_path(path: str, content: list, index: int) -> str:
    """
    Generate path for the selected subdirectory
    :param path: path of the current directory in string format
    :param content: list of subdirectories name in string format
    :param index: chosen integer option from menu
    :return: new_path: path of the desired subdirectory in string format
    """
    # Modify global variable
    global new_path
    
    # Generate path for specific subdirectory depends of the chosen option
    if index > 1:
        new_path = path + "/" + content[index - 3]
    else:
        temp_path = path.split("/")[:-1]
        new_path = "/".join(temp_path)
    
    # Return modified path variable
    return new_path


# Define function for menu creation
def generate_options(content: list, path: str) -> list:
    """
    Generate menu options for current directory
    :param content: list of subdirectories name in string format
    :param path: path of the current directory in string format
    :return: options: list of menu options
    """
    # Modify global variable
    global options

    # Generate menu options
    options.append("[0] Exit")
    options.append("[1] Back")
    options.append("[2] Save")
    new_option = ""

    # Append all subdirectories name to the options list
    for index, item in enumerate(content):
        current_file = path + "/" + content[index]
        if os.path.isdir(current_file):
            new_option = f"[{index + 3}] {item}"
            options.append(new_option)
    
    # Return modified options variable
    return options


# Define function for menu display
def display_options() -> None:
    """
    Generate menu options for current directory
    :param None
    :return: None
    """
    # Display in console all available options
    for item in options:
        print(item)


# Define function for menu reset
def reset_options() -> None:
    """
    Delete all current options
    :param None
    :return: None
    """
    # Modify global variable
    global options

    # Clear current content of the options variable
    options.clear()


# Define function for writing content in Excel file
def write_to_excel(file_name: str, file_path: str, path: str, content: list, target_row: int, target_column: int) -> None:
    """
    Write to Excel file the name of subdirectories from specified directory
    :param file_name: name of the Excel file in string format
    :param file_path: path of the directory where Excel file will be stored in string format
    :param path: path of the current directory in string format
    :param content: list of subdirectories name in string format
    :param target_row: desired row number from where start write data in integer format
    :param target_column: desired column number from where start write data in integer format
    :return: None
    """
    # Compile final file name and path
    final_file_name = file_name + ".xlsx"
    final_file_path = file_path + "/" + final_file_name
    # Create new Excel file or append to exiting one
    if os.path.isfile(final_file_path):
        xlsx_sheet = load_workbook(final_file_name)
    else:
        xlsx_sheet = Workbook()
    
    # Open Excel file for writing
    sheet = xlsx_sheet.active

    # Take the last row where data was written
    last_row = sheet.max_row
    last_row += target_row

    # Write data on specified column and starting from last empty row
    for item in range(len(content)):
        current_file = path + "/" + content[item]
        if os.path.isdir(current_file):
            sheet.cell(row=last_row, column=target_column, value=content[item])
            last_row += 1
    
    # Write delimiter between each directory content
    sheet.cell(row=last_row, column=target_column, value="*****")

    # Save Excel file
    xlsx_sheet.save(final_file_path)


# Define function for content copying
def copy_content(source_path: str) -> None:
    """
    Copy content from source to destination directory
    :param source_path: path of the current directory in string format
    :return: None
    """
    # Ask user to provide destination directory path
    destination_path = input("Provide destination path: ")

    # Copy whole content from source directory to destination directory
    distutils.dir_util.copy_tree(source_path, destination_path)


# Define function for display greetings
def display_greetings(state: str) -> None:
    """
    Display specific start and end message
    :param state: specify position of the message in string format
    :return: None
    """
    if state == "start":
        print("**************************")
        print("***** MIGRATION TOOL *****")
        print("**************************\n")
    else:
        print("*******************************")
        print("***** EXIT MIGRATION TOOL *****")
        print("*******************************\n")


# Run code as a script
if __name__ == "__main__":

    # Print initial message
    display_greetings("start")

    # Loop for existing path providing
    is_not_valid_path = True
    while is_not_valid_path:
        user_path = input("Provide directory path: ")
        # Assign existing path to the global variable
        new_path = user_path
        # Exit from loop if path exists
        if os.path.exists(user_path):
            is_not_valid_path = False
    
    # Read content and store it in a variable
    res_cont = read_content(new_path)
    # Generate options for menu with generated content
    generate_options(res_cont, new_path)
    # Display generated menu
    display_options()
    # Write to Excel file with specified name at specified location
    write_to_excel("MigrationSheet", "C:/Users/rmurz/Migration", new_path, res_cont, 1, 1)

    # Loop through user options
    is_continue = True
    while is_continue:
        # Take option from the user
        user_option = int(input("Choose option: "))

        # Exit from program if user press Exit option or press Back at the root level
        if user_option == 0 or (user_option == 1 and new_path == user_path):
            is_continue = False
            display_greetings("end")
            SystemExit

        # Display options if user go back
        elif user_option == 1:
            # Generate path for new selected directory
            generate_path(new_path, content, user_option)
            # Delete all previous options
            reset_options()
            # Read and store in a variable all subdirectories from current directory
            res_cont = read_content(new_path)
            # Generate options for menu with generated content
            generate_options(content, new_path)
            # Display generated menu
            display_options()


        # Copy the content of the current directory in specified directory
        elif user_option == 2:
            copy_content(new_path)

        # Display options and write to Excel file if user go forward
        else:
            # Generate path for new selected directory
            generate_path(new_path, content, user_option)
            # Delete all previous options
            reset_options()
            # Read and store in a variable all subdirectories from current directory
            res_cont = read_content(new_path)
            # Generate options for menu with generated content
            generate_options(content, new_path)
            # Display generated menu
            display_options()
            # Write to Excel file with specified name at specified location
            write_to_excel("MigrationSheet", "C:/Users/rmurz/Migration", new_path, res_cont, 1, 1)
