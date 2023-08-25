# Import necessary libraries
import os

# Import necessary modules
from office365.sharepoint.client_context import ClientContext
from openpyxl import Workbook, load_workbook

# Import internal utilities
from credentials import username, password, site_url

# Create connection to the SharePoint
context = ClientContext(site_url).with_user_credentials(username, password)


# Define function that retrieve all subfolders
def get_folder_data(folder_path: str) -> list:
    """
    Return a list with the names of subdirectories from provided URL
    :param folder_path: path of the directory in string format
    :return: folders_name: list with subdirectories name
    """
    # Get relative URL from SharePoint for given folder and load it
    library_root = context.web.get_folder_by_server_relative_path(folder_path)
    context.load(library_root)
    context.execute_query()

    # Retrieve all subfolders from provided folder
    folders = library_root.folders
    context.load(folders)
    context.execute_query()

    # Display the number of subfolders
    print("Number of folders:", len(folders))

    # Iterate through all subfolder, display and save the name
    folders_name = []
    for folder in folders:
        context.load(folder)
        context.execute_query()
        # print(folder.properties["ServerRelativeUrl"].split("/")[-1])
        folders_name.append(folder.properties["ServerRelativeUrl"].split("/")[-1])
    
    # Return a list of subfolders name
    return folders_name


# Define function that retrieve all files
def get_folder_files(folder_path: str) -> list:
    """
    Return a list with the names of files from folder
    :param folder_path: path of the directory in string format
    :return: folders_name: list with files name
    """
    # Get relative URL from SharePoint for given folder and load it
    library_root = context.web.get_folder_by_server_relative_path(folder_path)
    context.load(library_root)
    context.execute_query()
    
    # Retrieve all files from provided folder
    files = library_root.files
    context.load(files)
    context.execute_query()

    # Display the number of files
    print("Number of files:", len(files))

    # Iterate through all files, display and save the name
    files_name = []
    for file in files:
        context.load(file)
        context.execute_query()
        # print(file.properties["Name"])
        files_name.append(file.properties["Name"])
    
    # Return a list of files name
    return files_name


# Define function that download file from SharePoint
def download_file(file_url: str) -> None:
    """
    Download a file to local disk from SharePoint
    :param file_url: path of the SharePoint file in string format
    :return: None
    """
    # Retrieve file name from URL and build local path
    file_name = file_url.split("/")[-1]
    download_folder = os.getcwd().split("\\")
    download_file = "/".join(download_folder) + "/" + file_name

    # Open file and save content
    with open(download_file, "wb") as local_file:
        context.web.get_file_by_server_relative_url(file_url).download(local_file).execute_query()
    
    # Display a message of completion
    # print(f"File has been downloaded into: {download_file}")    


# Define function that upload file to SharePoint
def upload_file(file_name:str, folder_url: str) -> None:
    """
    Upload a file to SharePoint from local disk
    :param file_name: name of the file in string format
    :param folder_url: path of the SharePoint folder in string format
    :return: None
    """
    # Take folder from SharePoint based on URL
    folder = context.web.get_folder_by_server_relative_url(folder_url)

    # Take local working folder and retrieve local file
    download_folder = os.getcwd().split("\\")
    local_file = "/".join(download_folder) + "/" + file_name

    # Open file and save content
    with open(local_file, 'rb') as f:
        file = folder.files.upload(f).execute_query()

    # Display a message of completion
    # print(f"File has been uploaded into: {file.serverRelativeUrl}")


# Define function that copy the whole folder
def copy_folder(source_path: str, destination_path: str):
    """
    Copy the whole content of the directory from source directory to the target directory
    :param source_path: path of the source directory in string format
    :param destination_path: path of the destination directory in string format
    :return: None
    """
    # From raw URL obtain path to source directory and directory name
    raw_source_root = source_path.split("/")[:-1]
    source_root = "/".join(raw_source_root)
    source_folder = source_path.split("/")[-1]

    # From raw URL obtain path to destination directory and directory name
    raw_destination_root = destination_path.split("/")[:-1]
    destination_root = "/".join(raw_destination_root)
    destination_folder = destination_path.split("/")[-1]

    # Link obtained paths to the SharePoint
    folder_from = context.web.get_folder_by_server_relative_url(source_root).add(source_folder)
    folder_to = context.web.get_folder_by_server_relative_url(destination_root).add(destination_folder)

    # Copy directory
    folder_from.copy_to_using_path(folder_to).execute_query()

    # Display message of completion
    print(f"Folder has been copied from '{folder_from.serverRelativeUrl}' into '{folder_to.serverRelativeUrl}'")


# Define function that copy file
def copy_file(file_name: str, source_path: str, destination_path: str) -> None:
    """
    Copy file from source directory to the target directory
    :param temp_path: path of the local folder in string format
    :param file_name: name of the file in string format
    :param source_path: path of the SharePoint source file in string format
    :param destination_path: path of the SharePoint destination directory in string format
    :return: None
    """
    # Download the file to temporary location
    download_file(source_path)

    # Upload the file from temporary location
    upload_file(file_name, destination_path)

    # Delete file from local disk
    try:
        os.remove(file_name)
    except OSError:
        pass
    
    # Display message of completion
    print(f"File has been copied from '{source_path}' into '{destination_path}'")


# Define function for writing content in Excel file
def write_to_excel(file_name: str, file_path: str, content: list, target_row: int, target_column: int) -> None:
    """
    Write to Excel file the name of subdirectories from specified directory
    :param file_name: name of the Excel file in string format
    :param file_path: path of the directory where Excel file will be stored in string format
    :param content: list of subdirectories name in string format
    :param target_row: desired row number from where start write data in integer format
    :param target_column: desired column number from where start write data in integer format
    :return: None
    """
    # Compile final file name and path
    final_file_name = file_name + ".xlsx"
    final_file_path = file_path + "/" + final_file_name

    # Create new Excel file or append to exiting one
    if os.path.isfile(final_file_path):
        xlsx_sheet = load_workbook(final_file_name)
    else:
        xlsx_sheet = Workbook()
    
    # Open Excel file for writing
    sheet = xlsx_sheet.active

    # Take the last row where data was written
    last_row = sheet.max_row
    last_row += target_row

    # Write data on specified column and starting from last empty row
    for item in content:
        sheet.cell(row=last_row, column=target_column, value=item)
        last_row += 1
    
    # Write delimiter between each directory content
    sheet.cell(row=last_row, column=target_column, value="*****")

    # Save Excel file
    xlsx_sheet.save(final_file_path)

    # Display a message of completion
    print(f"Excel file {final_file_name} was updated successfuly.")


# Define function for reading content from Excel file
def read_from_excel(file_path: str, target_row: int, target_column: int) -> list:
    """
    Read from Excel file the name of subdirectories from specified directory
    :param file_path: path of the directory where Excel file is stored in string format
    :param target_row: desired row number from where start read data in integer format
    :param target_column: desired column number from where start read data in integer format
    :return: content: list of subdirectories name in string format
    """
    # Open Excel file for reading
    xlsx_sheet = load_workbook(file_path)
    sheet = xlsx_sheet.active

    # Take the last row where data was written
    last_row = sheet.max_row + target_row

    # Read content from one column and store it
    excel_content = []
    for index in range(last_row + 1):
        cell_content = sheet.cell(row=last_row, column=target_column).value
        excel_content.append(cell_content)
    
    # Return list of read values
    return excel_content


# Define function for writing item in Excel file
def log_to_excel(file_name: str, file_path: str, item: str, target_row: int, target_column: int) -> None:
    """
    Write to Excel file the name of subdirectories from specified directory
    :param file_name: name of the Excel file in string format
    :param file_path: path of the directory where Excel file will be stored in string format
    :param item: path of file or folder in string format
    :param target_row: desired row number from where start write data in integer format
    :param target_column: desired column number from where start write data in integer format
    :return: None
    """
    # Compile final file name and path
    final_file_name = file_name + ".xlsx"
    final_file_path = file_path + "/" + final_file_name

    # Create new Excel file or append to exiting one
    if os.path.isfile(final_file_path):
        xlsx_sheet = load_workbook(final_file_name)
    else:
        xlsx_sheet = Workbook()
    
    # Open Excel file for writing
    sheet = xlsx_sheet.active

    # Take the last row where data was written
    last_row = sheet.max_row
    last_row += target_row

    # Write data on specified column and starting from last empty row
    sheet.cell(row=last_row, column=target_column, value=item)
    last_row += 1
    
    # Write delimiter between each directory content
    sheet.cell(row=last_row, column=target_column, value="*****")

    # Save Excel file
    xlsx_sheet.save(final_file_path)

    # Display a message of completion
    print(f"Excel file {final_file_name} was updated successfuly.")


# Run code as a script
if __name__ == "__main__":

    # Define work scenario variables
    root_location = "/sites/<enterprise_site>/<enterprise_directory>"
    lvl_1_folder = ["/Folder_1", "/Folder_2", "/Folder_3", "/Folder_4", "/Folder_5"]
    lvl_1_1_folder = ["/Subfolder_1", "/Subfolder_1", "/Subfolder_1"]

    # Provide root work path
    for index in range(len(lvl_1_1_folder)):
        folder_path = root_location + lvl_1_folder[0] + lvl_1_1_folder[index]

        # Get subfolders name from folder
        folders_name = get_folder_data(folder_path)
        write_to_excel("/Folders_Path", os.getcwd(), folders_name, 1, 1)
